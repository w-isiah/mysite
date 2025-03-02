import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Marks Template"

# Add data to the first sheet
ws1.append(["reg no", "Term", "School", "Marks", "Assessment Type"])
ws1.append(["24/U/PPD/05949/PD", "Semester_1 Year 2025", "St. Bruno Sserunkuuma's SS", 98, "modulate"])

# Create a second sheet for drop-down data
ws2 = wb.create_sheet("drop_down data")

# Add data to the second sheet
ws2.append(["Terms", "asssessment type", "", "schools"])
ws2.append(["Semester_2 Year 2024", "modulate", "", "St. Bruno Sserunkuuma's SS"])
ws2.append(["Semester_1 Year 2025", "manual", "", ""])

# Extract drop-down options from the "drop_down data" sheet
def get_column_values(sheet, col_idx, min_row, max_row):
    values = []
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=col_idx, max_col=col_idx, values_only=True):
        if row[0] is not None:  # Check if the cell is not empty
            values.append(row[0])
    return values

terms = get_column_values(ws2, col_idx=1, min_row=2, max_row=3)
assessment_types = get_column_values(ws2, col_idx=2, min_row=2, max_row=3)
schools = get_column_values(ws2, col_idx=4, min_row=2, max_row=3)

# Create data validation for Term
dv_term = DataValidation(type="list", formula1=f'"{",".join(terms)}"', allow_blank=True)
ws1.add_data_validation(dv_term)
dv_term.add("B2:B1048576")  # Apply to all rows in the Term column

# Create data validation for Assessment Type
dv_assessment = DataValidation(type="list", formula1=f'"{",".join(assessment_types)}"', allow_blank=True)
ws1.add_data_validation(dv_assessment)
dv_assessment.add("E2:E1048576")  # Apply to all rows in the Assessment Type column

# Create data validation for School
dv_school = DataValidation(type="list", formula1=f'"{",".join(schools)}"', allow_blank=True)
ws1.add_data_validation(dv_school)
dv_school.add("C2:C1048576")  # Apply to all rows in the School column

# Save the workbook
wb.save("marks_template.xlsx")